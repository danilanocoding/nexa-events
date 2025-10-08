#export.py
from flask import Blueprint, send_file, jsonify
from models import db, Event, Participant
import openpyxl
from io import BytesIO

export_bp = Blueprint('export', __name__)

@export_bp.route('/export/<event_code>', methods=['GET'])
def export_participants(event_code):
    event = Event.query.filter_by(event_code=event_code).first()
    if not event:
        return jsonify({'status': 'error', 'message': 'Мероприятие не найдено'}), 404

    participants = Participant.query.filter_by(event_id=event.id).all()

    wb = openpyxl.Workbook()

    ws_all = wb.active
    ws_all.title = "Все участники"
    ws_present = wb.create_sheet(title="Пришли")
    ws_absent = wb.create_sheet(title="Не пришли")

    headers = ["№", "ФИО", "Академическая группа", "Статус"]

    def fill_sheet(ws, rows):
        ws.append(headers)
        max_lens = [len(h) for h in headers]
        for r in rows:
            ws.append(r)
            for i, val in enumerate(r):
                l = len(str(val)) if val is not None else 0
                if l > max_lens[i]:
                    max_lens[i] = l
        for i, ml in enumerate(max_lens, start=1):
            col = ws.cell(row=1, column=i).column_letter
            if i == 1:
                ws.column_dimensions[col].width = min(max(3, ml + 1), 6)  # narrow № column
            elif headers[i-1] == 'Академическая группа':
                ws.column_dimensions[col].width = min(max(10, ml + 2), 25)
            elif headers[i-1] == 'ФИО':
                ws.column_dimensions[col].width = min(max(20, ml + 2), 50)
            else:
                ws.column_dimensions[col].width = min(max(10, ml + 2), 30)

    rows_all = []
    rows_present = []
    rows_absent = []
    for idx, p in enumerate(participants, start=1):
        row = [idx, p.full_name or '', p.group or '', 'Присутствовал' if p.visited else 'Не явился']
        rows_all.append(row)
        if p.visited:
            rows_present.append(row)
        else:
            rows_absent.append(row)

    fill_sheet(ws_all, rows_all)
    fill_sheet(ws_present, rows_present)
    fill_sheet(ws_absent, rows_absent)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"{event.name.replace(' ', '_')}_участники.xlsx"
    return send_file(
        stream,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
